#!/usr/bin/env python3
"""
Excel Dashboard Generator
Creates a professional Excel dashboard with charts, KPIs, and formatting
"""

from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import sys

def create_excel_dashboard(data):
    """Create Excel dashboard with charts and formatting"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create Dashboard sheet
    dashboard = wb.create_sheet('Dashboard', 0)
    
    # Create data sheets
    create_kpi_sheet(wb, data)
    create_product_sheet(wb, data)
    create_branch_sheet(wb, data)
    create_monthly_sheet(wb, data)
    create_customer_sheet(wb, data)
    create_payment_sheet(wb, data)
    
    # Build dashboard with charts
    build_dashboard(dashboard, wb, data)
    
    return wb

def create_kpi_sheet(wb, data):
    """Create KPI data sheet"""
    sheet = wb.create_sheet('KPI_Data')
    kpis = data['kpis']
    
    # Headers
    sheet['A1'] = 'Metric'
    sheet['B1'] = 'Value'
    
    # Data
    sheet['A2'] = 'Total Sales'
    sheet['B2'] = kpis['totalSales']
    
    sheet['A3'] = 'Total Transactions'
    sheet['B3'] = kpis['totalTransactions']
    
    sheet['A4'] = 'Avg Transaction Value'
    sheet['B4'] = kpis['averageTransactionValue']
    
    sheet['A5'] = 'Avg Rating'
    sheet['B5'] = kpis['avgRating']
    
    # Format
    for row in sheet['A1:B1']:
        for cell in row:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=12)

def create_product_sheet(wb, data):
    """Create product line data sheet"""
    sheet = wb.create_sheet('Product_Data')
    products = data['dashboardData']['productLineSales']
    
    # Headers
    headers = ['Product Line', 'Sales', 'Transactions', 'Avg Value']
    for idx, header in enumerate(headers, 1):
        cell = sheet.cell(1, idx, header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Data
    for idx, product in enumerate(products, 2):
        sheet.cell(idx, 1, product['name'])
        sheet.cell(idx, 2, product['value'])
        sheet.cell(idx, 3, product['transactions'])
        sheet.cell(idx, 4, product['avgValue'])
    
    # Auto-size columns
    for col in range(1, 5):
        sheet.column_dimensions[get_column_letter(col)].width = 25

def create_branch_sheet(wb, data):
    """Create branch data sheet"""
    sheet = wb.create_sheet('Branch_Data')
    branches = data['dashboardData']['branchData']
    
    # Headers
    headers = ['Branch', 'City', 'Sales', 'Transactions', 'Rating']
    for idx, header in enumerate(headers, 1):
        cell = sheet.cell(1, idx, header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # Data
    for idx, branch in enumerate(branches, 2):
        sheet.cell(idx, 1, branch['branch'])
        sheet.cell(idx, 2, branch['city'])
        sheet.cell(idx, 3, branch['sales'])
        sheet.cell(idx, 4, branch['transactions'])
        sheet.cell(idx, 5, branch['rating'])
    
    # Auto-size columns
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 18

def create_monthly_sheet(wb, data):
    """Create monthly data sheet"""
    sheet = wb.create_sheet('Monthly_Data')
    monthly = data['dashboardData']['monthlyData']
    
    # Headers
    headers = ['Month', 'Sales', 'Transactions']
    for idx, header in enumerate(headers, 1):
        cell = sheet.cell(1, idx, header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Data
    for idx, month in enumerate(monthly, 2):
        sheet.cell(idx, 1, month['month'])
        sheet.cell(idx, 2, month['sales'])
        sheet.cell(idx, 3, month['transactions'])
    
    # Auto-size columns
    for col in range(1, 4):
        sheet.column_dimensions[get_column_letter(col)].width = 20

def create_customer_sheet(wb, data):
    """Create customer data sheet"""
    sheet = wb.create_sheet('Customer_Data')
    customers = data['dashboardData']['customerData']
    
    # Headers
    headers = ['Type', 'Gender', 'Sales', 'Transactions', 'Rating']
    for idx, header in enumerate(headers, 1):
        cell = sheet.cell(1, idx, header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Data
    for idx, customer in enumerate(customers, 2):
        sheet.cell(idx, 1, customer['type'])
        sheet.cell(idx, 2, customer['gender'])
        sheet.cell(idx, 3, customer['sales'])
        sheet.cell(idx, 4, customer['transactions'])
        sheet.cell(idx, 5, customer['rating'])
    
    # Auto-size columns
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 18

def create_payment_sheet(wb, data):
    """Create payment data sheet"""
    sheet = wb.create_sheet('Payment_Data')
    payments = data['dashboardData']['paymentData']
    
    # Headers
    headers = ['Payment Method', 'Sales', 'Transactions']
    for idx, header in enumerate(headers, 1):
        cell = sheet.cell(1, idx, header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='A5A5A5', end_color='A5A5A5', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Data
    for idx, payment in enumerate(payments, 2):
        sheet.cell(idx, 1, payment['method'])
        sheet.cell(idx, 2, payment['sales'])
        sheet.cell(idx, 3, payment['transactions'])
    
    # Auto-size columns
    for col in range(1, 4):
        sheet.column_dimensions[get_column_letter(col)].width = 22

def build_dashboard(sheet, wb, data):
    """Build the main dashboard with KPIs and charts"""
    kpis = data['kpis']
    
    # Title
    sheet.merge_cells('A1:L1')
    title = sheet['A1']
    title.value = 'SUPERMARKET SALES DASHBOARD'
    title.font = Font(bold=True, size=20, color='FFFFFF')
    title.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 35
    
    # Subtitle
    sheet.merge_cells('A2:L2')
    subtitle = sheet['A2']
    subtitle.value = 'Comprehensive Sales Analytics and Performance Insights'
    subtitle.font = Font(size=12, italic=True, color='7F7F7F')
    subtitle.alignment = Alignment(horizontal='center')
    
    # KPI Cards
    kpi_row = 4
    kpi_data = [
        ('Total Sales', f"₹{kpis['totalSales']:,.0f}", '70AD47'),
        ('Total Transactions', f"{kpis['totalTransactions']:,}", '5B9BD5'),
        ('Avg Transaction', f"₹{kpis['averageTransactionValue']:,.0f}", 'FFC000'),
        ('Avg Rating', f"{kpis['avgRating']:.1f}/10", 'ED7D31')
    ]
    
    col_offset = 1
    for title, value, color in kpi_data:
        # Merge cells for KPI card
        sheet.merge_cells(start_row=kpi_row, start_column=col_offset, 
                         end_row=kpi_row, end_column=col_offset+2)
        sheet.merge_cells(start_row=kpi_row+1, start_column=col_offset, 
                         end_row=kpi_row+1, end_column=col_offset+2)
        
        # Title
        title_cell = sheet.cell(kpi_row, col_offset)
        title_cell.value = title
        title_cell.font = Font(size=10, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        title_cell.font = Font(bold=True, color='FFFFFF')
        
        # Value
        value_cell = sheet.cell(kpi_row+1, col_offset)
        value_cell.value = value
        value_cell.font = Font(size=16, bold=True, color=color)
        value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        sheet.row_dimensions[kpi_row+1].height = 30
        
        col_offset += 3
    
    # Add charts
    chart_row = 8
    
    # Product Line Pie Chart
    add_product_pie_chart(sheet, wb, chart_row, 1)
    
    # Monthly Trend Line Chart
    add_monthly_line_chart(sheet, wb, chart_row, 7)
    
    # Branch Bar Chart
    add_branch_bar_chart(sheet, wb, chart_row + 18, 1)
    
    # Payment Pie Chart
    add_payment_pie_chart(sheet, wb, chart_row + 18, 7)
    
    # Key Insights
    add_insights(sheet, chart_row + 36)
    
    # Adjust column widths
    for col in range(1, 13):
        sheet.column_dimensions[get_column_letter(col)].width = 12

def add_product_pie_chart(sheet, wb, row, col):
    """Add product line pie chart"""
    product_sheet = wb['Product_Data']
    
    chart = PieChart()
    labels = Reference(product_sheet, min_col=1, min_row=2, max_row=7)
    data = Reference(product_sheet, min_col=2, min_row=1, max_row=7)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Sales by Product Line"
    chart.height = 12
    chart.width = 18
    
    sheet.add_chart(chart, f'{get_column_letter(col)}{row}')

def add_monthly_line_chart(sheet, wb, row, col):
    """Add monthly trend line chart"""
    monthly_sheet = wb['Monthly_Data']
    
    chart = LineChart()
    chart.title = "Monthly Sales Trend"
    chart.y_axis.title = "Sales (₹)"
    chart.x_axis.title = "Month"
    
    data = Reference(monthly_sheet, min_col=2, min_row=1, max_row=4)
    cats = Reference(monthly_sheet, min_col=1, min_row=2, max_row=4)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18
    
    sheet.add_chart(chart, f'{get_column_letter(col)}{row}')

def add_branch_bar_chart(sheet, wb, row, col):
    """Add branch performance bar chart"""
    branch_sheet = wb['Branch_Data']
    
    chart = BarChart()
    chart.title = "Branch Performance"
    chart.y_axis.title = "Sales (₹)"
    
    data = Reference(branch_sheet, min_col=3, min_row=1, max_row=4)
    cats = Reference(branch_sheet, min_col=2, min_row=2, max_row=4)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18
    
    sheet.add_chart(chart, f'{get_column_letter(col)}{row}')

def add_payment_pie_chart(sheet, wb, row, col):
    """Add payment method pie chart"""
    payment_sheet = wb['Payment_Data']
    
    chart = PieChart()
    labels = Reference(payment_sheet, min_col=1, min_row=2, max_row=4)
    data = Reference(payment_sheet, min_col=2, min_row=1, max_row=4)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Payment Method Distribution"
    chart.height = 12
    chart.width = 18
    
    sheet.add_chart(chart, f'{get_column_letter(col)}{row}')

def add_insights(sheet, row):
    """Add key insights section"""
    sheet.merge_cells(f'A{row}:L{row}')
    title = sheet.cell(row, 1)
    title.value = 'KEY INSIGHTS'
    title.font = Font(bold=True, size=14, color='FFFFFF')
    title.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[row].height = 25
    
    insights = [
        '• Food & Beverages leads in total sales (₹56,145)',
        '• Naypyitaw branch has highest sales (₹1,10,569) and rating (7.1/10)',
        '• March showed strong recovery with ₹1,09,456 in sales',
        '• Cash remains the preferred payment method (34.7% of total sales)'
    ]
    
    for idx, insight in enumerate(insights, 1):
        sheet.merge_cells(f'A{row+idx}:L{row+idx}')
        cell = sheet.cell(row+idx, 1)
        cell.value = insight
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        sheet.row_dimensions[row+idx].height = 20

if __name__ == '__main__':
    # Read JSON data from stdin
    input_data = json.loads(sys.stdin.read())
    
    # Create workbook
    wb = create_excel_dashboard(input_data)
    
    # Save to file
    output_file = input_data.get('outputFile', 'dashboard.xlsx')
    wb.save(output_file)
    
    print(json.dumps({'success': True, 'file': output_file}))